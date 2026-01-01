"""
Excel parser service.

This module handles parsing Excel files (.xlsx and .xls) into internal data structures.
"""

from pathlib import Path

import openpyxl
import openpyxl.utils
import xlrd
from openpyxl.cell.cell import Cell as OpenpyxlCell
from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

from src.lib.exceptions import CorruptedFile, ConversionFailed
from src.lib.logging import get_logger
from src.models import FileFormat
from src.models.excel import ExcelCell, ExcelSheet, ExcelWorkbook
from src.services.converter.color_utils import get_color_from_color_object, hex_to_rgb

logger = get_logger(__name__)


def _parse_xlsx_cell(cell: OpenpyxlCell, row_idx: int, col_idx: int) -> ExcelCell:
    """
    Parse an openpyxl cell into ExcelCell.

    Args:
        cell: openpyxl Cell object
        row_idx: Row index (0-based)
        col_idx: Column index (0-based)

    Returns:
        ExcelCell with formatting information
    """
    # Get cell value
    value = cell.value

    # Get font properties
    font = cell.font if cell.font else Font()
    font_name = font.name or "Arial"
    font_size = float(font.size) if font.size else 11.0
    font_bold = font.bold or False
    font_italic = font.italic or False

    # Get font color
    font_color = "000000"
    if font.color and font.color.rgb:
        font_color = hex_to_rgb(str(font.color.rgb))

    # Get background color
    bg_color = None
    if cell.fill and hasattr(cell.fill, "patternType") and cell.fill.patternType:
        color_attr = None
        if hasattr(cell.fill, "fgColor") and cell.fill.fgColor:
            color_attr = cell.fill.fgColor
        elif hasattr(cell.fill, "start_color") and cell.fill.start_color:
            color_attr = cell.fill.start_color
        bg_color = get_color_from_color_object(color_attr)

    # Get alignment
    alignment = cell.alignment if cell.alignment else Alignment()
    alignment_h = alignment.horizontal or "left"
    alignment_v = alignment.vertical or "top"
    wrap_text = alignment.wrap_text or False

    # Get text rotation
    text_rotation = 0
    if alignment.textRotation:
        text_rotation = alignment.textRotation

    # Get number format
    number_format = cell.number_format if cell.number_format else "General"

    # Check for borders and capture individual side styles
    has_border = False
    border_style = "thin"  # Default for backward compatibility
    border_left = None
    border_right = None
    border_top = None
    border_bottom = None

    if cell.border:
        border = cell.border
        # Collect all border styles present
        styles = []

        if border.left and border.left.style:
            has_border = True
            border_left = border.left.style
            styles.append(border.left.style)
        if border.right and border.right.style:
            has_border = True
            border_right = border.right.style
            styles.append(border.right.style)
        if border.top and border.top.style:
            has_border = True
            border_top = border.top.style
            styles.append(border.top.style)
        if border.bottom and border.bottom.style:
            has_border = True
            border_bottom = border.bottom.style
            styles.append(border.bottom.style)

        # Determine the most prominent border style for backward compatibility
        # Priority: medium > thick > dashed > dotted > thin
        if styles:
            if "medium" in styles:
                border_style = "medium"
            elif "thick" in styles:
                border_style = "thick"
            elif "dashed" in styles:
                border_style = "dashed"
            elif "dotted" in styles:
                border_style = "dotted"
            else:
                border_style = "thin"

    return ExcelCell(
        value=value,
        row=row_idx,
        column=col_idx,
        font_name=font_name,
        font_size=font_size,
        font_bold=font_bold,
        font_italic=font_italic,
        font_color=font_color,
        bg_color=bg_color,
        alignment_horizontal=alignment_h,
        alignment_vertical=alignment_v,
        text_rotation=text_rotation,
        wrap_text=wrap_text,
        has_border=has_border,
        border_style=border_style,
        border_left=border_left,
        border_right=border_right,
        border_top=border_top,
        border_bottom=border_bottom,
        number_format=number_format,
    )


def _parse_xlsx(file_path: Path) -> ExcelWorkbook:
    """
    Parse .xlsx file using openpyxl.

    Args:
        file_path: Path to .xlsx file

    Returns:
        ExcelWorkbook object

    Raises:
        CorruptedFile: If file cannot be parsed
    """
    try:
        wb: OpenpyxlWorkbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error(f"Failed to load .xlsx file: {e}")
        raise CorruptedFile(f"Failed to parse Excel file: {str(e)}")

    sheets: list[ExcelSheet] = []

    for ws in wb.worksheets:
        cells: list[ExcelCell] = []
        max_row = ws.max_row or 0
        max_column = ws.max_column or 0

        # Build merged cells map: (row, col) -> (row_span, col_span)
        merged_cells_map: dict[tuple[int, int], tuple[int, int]] = {}
        for merged_range in ws.merged_cells.ranges:
            min_row, min_col = merged_range.min_row - 1, merged_range.min_col - 1
            max_row_range, max_col_range = (
                merged_range.max_row - 1,
                merged_range.max_col - 1,
            )
            row_span = max_row_range - min_row + 1
            col_span = max_col_range - min_col + 1
            merged_cells_map[(min_row, min_col)] = (row_span, col_span)

        # Capture column widths (Excel units, default ~8.43)
        column_widths: dict[int, float] = {}
        for col_idx in range(max_column):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            width = ws.column_dimensions[col_letter].width
            if width is not None:
                column_widths[col_idx] = width

        # Capture row heights (points, default ~15)
        row_heights: dict[int, float] = {}
        for row_idx in range(max_row):
            height = ws.row_dimensions[row_idx + 1].height
            if height is not None:
                row_heights[row_idx] = height

        # Parse all cells in the sheet
        for row_idx, row in enumerate(ws.iter_rows()):
            for col_idx, cell in enumerate(row):
                # Parse cells that have content OR borders
                has_content = cell.value is not None
                has_border = False
                if cell.border:
                    has_border = any(
                        [
                            cell.border.left and cell.border.left.style,
                            cell.border.right and cell.border.right.style,
                            cell.border.top and cell.border.top.style,
                            cell.border.bottom and cell.border.bottom.style,
                        ]
                    )

                if has_content or has_border:
                    excel_cell = _parse_xlsx_cell(cell, row_idx, col_idx)

                    # Check if this cell is part of a merged range
                    if (row_idx, col_idx) in merged_cells_map:
                        row_span, col_span = merged_cells_map[(row_idx, col_idx)]
                        excel_cell.row_span = row_span
                        excel_cell.col_span = col_span

                    cells.append(excel_cell)

        sheet = ExcelSheet(
            name=ws.title,
            cells=cells,
            max_row=max_row,
            max_column=max_column,
            column_widths=column_widths,
            row_heights=row_heights,
        )
        sheets.append(sheet)

    return ExcelWorkbook(
        sheets=sheets,
        filename=file_path.name,
        format=FileFormat.XLSX,
    )


def _parse_xls(file_path: Path) -> ExcelWorkbook:
    """
    Parse .xls file using xlrd.

    Args:
        file_path: Path to .xls file

    Returns:
        ExcelWorkbook object

    Raises:
        CorruptedFile: If file cannot be parsed
    """
    try:
        wb = xlrd.open_workbook(str(file_path), formatting_info=True)
    except Exception as e:
        logger.error(f"Failed to load .xls file: {e}")
        raise CorruptedFile(f"Failed to parse Excel file: {str(e)}")

    sheets: list[ExcelSheet] = []

    for sheet_idx in range(wb.nsheets):
        ws = wb.sheet_by_index(sheet_idx)
        cells: list[ExcelCell] = []

        # Parse all cells in the sheet
        for row_idx in range(ws.nrows):
            for col_idx in range(ws.ncols):
                cell = ws.cell(row_idx, col_idx)

                if cell.value:
                    # Basic parsing for .xls (limited formatting support)
                    excel_cell = ExcelCell(
                        value=cell.value,
                        row=row_idx,
                        column=col_idx,
                        font_name="Arial",
                        font_size=11.0,
                        font_bold=False,
                        font_italic=False,
                        font_color="000000",
                        bg_color=None,
                        alignment_horizontal="left",
                        alignment_vertical="top",
                        has_border=False,
                    )
                    cells.append(excel_cell)

        sheet = ExcelSheet(
            name=ws.name,
            cells=cells,
            max_row=ws.nrows,
            max_column=ws.ncols,
        )
        sheets.append(sheet)

    return ExcelWorkbook(
        sheets=sheets,
        filename=file_path.name,
        format=FileFormat.XLS,
    )


class ExcelParser:
    """Service for parsing Excel files."""

    def parse(self, file_path: Path, file_format: FileFormat) -> ExcelWorkbook:
        """
        Parse an Excel file into internal data structure.

        Args:
            file_path: Path to Excel file
            file_format: File format (XLSX or XLS)

        Returns:
            ExcelWorkbook object

        Raises:
            CorruptedFile: If file cannot be parsed
            ConversionFailed: If parsing fails for other reasons
        """
        logger.info(f"Parsing Excel file: {file_path.name} (format: {file_format})")

        try:
            if file_format == FileFormat.XLSX:
                workbook = _parse_xlsx(file_path)
            elif file_format == FileFormat.XLS:
                workbook = _parse_xls(file_path)
            else:
                raise ConversionFailed(f"Unsupported file format: {file_format}")

            logger.info(
                f"Successfully parsed {len(workbook.sheets)} sheet(s) "
                f"from {file_path.name}"
            )

            return workbook

        except CorruptedFile:
            raise
        except Exception as e:
            logger.error(f"Excel parsing failed: {e}")
            raise ConversionFailed(f"Failed to parse Excel file: {str(e)}")


# Singleton instance
_excel_parser: ExcelParser | None = None


def get_excel_parser() -> ExcelParser:
    """
    Get the Excel parser service instance.

    Returns:
        ExcelParser instance
    """
    global _excel_parser
    if _excel_parser is None:
        _excel_parser = ExcelParser()
    return _excel_parser


__all__ = [
    "ExcelParser",
    "get_excel_parser",
]
